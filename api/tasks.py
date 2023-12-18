from celery import shared_task
from openpyxl import load_workbook
from api.models import Customer, Loan
import pandas as pd


@shared_task
def ingest_data():
    # # Load customer data
    # customer_data = pd.read_excel("data/customer_data.xlsx")
    # for _, row in customer_data.iterrows():
    #     Customer.objects.create(**row)
    # # Load loan data
    # loan_data = pd.read_excel("data/loan_data.xlsx")
    # for _, row in loan_data.iterrows():
    #     customer_id = row['customer_id']
    #     try :
    #         customer = Customer.objects.get(customer_id=customer_id)
    #         Loan.objects.create(customer=customer, **row)
    #     except Customer.DoesNotExist:
    #         print(f'{customer_id} does not exists')

    # Load customer data
    customer_wb = load_workbook('data/customer_data.xlsx')
    customer_sheet = customer_wb.active

    for row in customer_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
                # print(row[0], type(row[0]))
                # print(row[3], type(row[3]))
                # print(row[4], type(row[4]))
                # print(row[5], type(row[5]))
                # print(row[6], type(row[6]))
                customer = Customer.objects.get_or_create(
                    customer_id=row[0],
                    first_name=row[1],
                    last_name=row[2],
                    age=row[3],
                    phone_number=row[4],
                    monthly_salary=row[5],
                    approved_limit=row[6],
                )
                print(customer)
        else:
            print(f'Customer {row[0]} cannot be saved')

    # Load loan data
    loan_wb = load_workbook('data/loan_data.xlsx')
    loan_sheet = loan_wb.active

    for row in loan_sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            try :
                customer_ID = Customer.objects.get(customer_id=row[0])
                loan = Loan.objects.get_or_create(
                    customer=customer_ID,
                    loan_id=row[1],
                    loan_amount=row[2],
                    tenure=row[3],
                    interest_rate=row[4],
                    monthly_repayment=row[5],
                    emis_paid_on_time=row[6],
                    start_date=row[7],
                    end_date=row[8],
                )
                print(loan)
            except Customer.DoesNotExist:
                print(f'Customer ID:{row[0]} does not exists, hence loan data cannot be uploaded')
        else:
             print(f'Loan ID {row[1]} cannot be saved')
        